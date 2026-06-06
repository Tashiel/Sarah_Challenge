import csv
import json
import os
from pathlib import Path
from typing import Any, Dict, List

from twitter_client import fetch_twitter_posts


Post = Dict[str, Any]


def ingest_file(path: str) -> List[Post]:
    source_path = Path(path)
    if not source_path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    if source_path.suffix.lower() == ".json":
        with source_path.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
        if isinstance(data, dict):
            data = data.get("posts", [])
        if not isinstance(data, list):
            raise ValueError("JSON file must contain a list of posts or an object with a `posts` list.")
        return data

    if source_path.suffix.lower() == ".csv":
        with source_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            return [row for row in reader]

    if source_path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required to read Excel files. Install it with `pip install openpyxl`."
            ) from exc

        workbook = load_workbook(source_path, read_only=True, data_only=True)
        posts: List[Post] = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            if not any(headers):
                continue

            for row_index, row in enumerate(rows[1:], start=2):
                if all(cell is None for cell in row):
                    continue

                row_data = {
                    headers[col_index]: row[col_index]
                    for col_index in range(min(len(headers), len(row)))
                    if headers[col_index]
                }

                text_value = (
                    row_data.get("text")
                    or row_data.get("Text")
                    or row_data.get("description")
                    or row_data.get("Description")
                    or row_data.get("Name")
                    or row_data.get("name")
                )
                if not text_value:
                    text_value = " ".join(
                        str(value).strip() for value in row if value is not None
                    )

                post_id = str(
                    row_data.get("id")
                    or row_data.get("ID")
                    or row_data.get("Uid")
                    or row_data.get("UID")
                    or f"{sheet_name}-{row_index}"
                )

                posts.append(
                    {
                        "id": post_id,
                        "text": str(text_value),
                        "author": row_data.get("author")
                        or row_data.get("Author")
                        or sheet_name,
                        "created_at": row_data.get("created_at")
                        or row_data.get("Created At")
                        or row_data.get("created_at_utc"),
                        "source": source_path.name,
                    }
                )

        return posts

    raise ValueError("Unsupported file type. Provide a .json, .csv, or .xlsx file.")


def ingest_twitter(query: str, limit: int = 50, bearer_token: str | None = None) -> List[Post]:
    if not bearer_token:
        print("TWITTER_BEARER_TOKEN is not set. Falling back to sample data for the POC.")
        sample_path = Path(__file__).parent / "sample_data" / "posts.json"
        if sample_path.exists():
            return ingest_file(str(sample_path))[:limit]
        return []
    return fetch_twitter_posts(query=query, limit=limit, bearer_token=bearer_token)
